import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import json
from datetime import datetime

def fetch_swagger_docs(swagger_url):
    """Fetch Swagger/OpenAPI documentation from the API"""
    try:
        print(f"Fetching Swagger docs from {swagger_url}...")
        response = requests.get(swagger_url)
        response.raise_for_status()
        print("Successfully fetched Swagger documentation")
        return response.json()
    except requests.exceptions.RequestException as e:
        print(f"Error fetching Swagger docs: {e}")
        return None

def get_model_details(swagger_data, model_ref):
    """Extract complete details about a model including examples and schema"""
    if not model_ref or not isinstance(model_ref, str):
        return None
    
    model_name = model_ref.split('/')[-1]
    model_def = swagger_data.get('definitions', {}).get(model_name, {})
    
    # Get example value
    example = model_def.get('example', '')
    if isinstance(example, (dict, list)):
        example = json.dumps(example, indent=2)
    
    # Get full schema with descriptions
    properties = model_def.get('properties', {})
    schema_info = []
    required_fields = model_def.get('required', [])
    
    for prop_name, prop_def in properties.items():
        prop_type = prop_def.get('type', 'object')
        if 'items' in prop_def and '$ref' in prop_def['items']:
            prop_type = f"array[{prop_def['items']['$ref'].split('/')[-1]}]"
        elif '$ref' in prop_def:
            prop_type = prop_def['$ref'].split('/')[-1]
        
        prop_desc = prop_def.get('description', '')
        required = " (required)" if prop_name in required_fields else ""
        schema_info.append({
            'name': prop_name,
            'type': prop_type,
            'description': prop_desc,
            'required': required,
            'example': prop_def.get('example', '')
        })
    
    return {
        'name': model_name,
        'example': str(example),
        'schema': schema_info,
        'description': model_def.get('description', '')
    }

def format_cell(cell, value, wrap_text=True):
    """Format a cell with consistent styling"""
    cell.value = value
    cell.alignment = Alignment(wrap_text=wrap_text, vertical='top')
    cell.border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    return cell

def create_excel_sheet(ws, title, headers, data):
    """Create a formatted Excel sheet with data"""
    # Set sheet title
    ws.title = title
    
    # Create header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    
    # Write headers
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, horizontal='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Write data
    for row_num, row_data in enumerate(data, 2):
        for col_num, cell_value in enumerate(row_data, 1):
            format_cell(ws.cell(row=row_num, column=col_num), cell_value)
    
    # Adjust column widths
    for col_num in range(1, len(headers) + 1):
        max_length = 0
        column_letter = get_column_letter(col_num)
        
        for cell in ws[column_letter]:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        
        adjusted_width = min((max_length + 2) * 1.2, 50)  # Cap at 50
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Freeze headers
    ws.freeze_panes = 'A2'
    
    return ws

def parse_swagger_to_excel(swagger_data, output_file):
    """Parse Swagger data and save to Excel with comprehensive details"""
    if not swagger_data:
        print("No Swagger data to process")
        return
    
    print("Creating Excel workbook...")
    wb = openpyxl.Workbook()
    
    # Remove default sheet if not needed
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Prepare endpoints data
    endpoints_data = []
    for path, path_item in swagger_data.get('paths', {}).items():
        for method, operation in path_item.items():
            if method.lower() not in ['get', 'post', 'put', 'delete', 'patch']:
                continue
            
            # Collect parameters
            parameters = []
            for param in operation.get('parameters', []):
                param_info = f"{param.get('name', '')} ({param.get('in', '')}): {param.get('type', 'object')}"
                if 'description' in param:
                    param_info += f" - {param['description']}"
                parameters.append(param_info)
            
            # Collect responses and models
            responses = []
            response_models = []
            for status_code, response in operation.get('responses', {}).items():
                resp_info = f"{status_code}: {response.get('description', '')}"
                responses.append(resp_info)
                
                if 'schema' in response:
                    ref = response['schema'].get('$ref', '')
                    if ref:
                        model_name = ref.split('/')[-1]
                        response_models.append(model_name)
            
            endpoints_data.append([
                method.upper(),
                path,
                operation.get('summary', ''),
                operation.get('operationId', ''),
                ", ".join(operation.get('consumes', [])),
                ", ".join(operation.get('produces', [])),
                "\n".join(parameters),
                "\n".join(responses),
                ", ".join(response_models),
                ", ".join(operation.get('tags', []))
            ])
    
    # Create endpoints sheet
    endpoint_headers = [
        "HTTP Method", "Path", "Summary", "Operation ID", 
        "Consumes", "Produces", "Parameters", 
        "Responses", "Response Models", "Tags"
    ]
    create_excel_sheet(wb.create_sheet("Endpoints"), "Endpoints", endpoint_headers, endpoints_data)
    
    # Prepare models data
    models_data = []
    model_examples_data = []
    all_models = swagger_data.get('definitions', {})
    
    for model_name, model_def in all_models.items():
        model_details = get_model_details(swagger_data, f"#/definitions/{model_name}")
        
        # Add to model examples sheet
        model_examples_data.append([
            model_details['name'],
            model_details['example'],
            model_details['description']
        ])
        
        # Add detailed properties to models sheet
        if not model_details['schema']:
            models_data.append([model_name, '', '', '', '', '', ''])
            continue
            
        for prop in model_details['schema']:
            example = prop['example']
            if isinstance(example, (dict, list)):
                example = json.dumps(example, indent=2)
            
            models_data.append([
                model_name,
                prop['name'],
                prop['type'],
                prop['description'],
                "Yes" if prop['required'] else "No",
                str(example),
                ""
            ])
    
    # Create models sheet
    model_headers = [
        "Model Name", "Property", "Type", "Description", 
        "Required", "Example", "Reference"
    ]
    create_excel_sheet(wb.create_sheet("Models"), "Models", model_headers, models_data)
    
    # Create model examples sheet
    example_headers = [
        "Model Name", "Example Value", "Description"
    ]
    create_excel_sheet(wb.create_sheet("Model Examples"), "Model Examples", example_headers, model_examples_data)
    
    # Create schema details sheet
    schema_data = []
    for model_name, model_def in all_models.items():
        model_details = get_model_details(swagger_data, f"#/definitions/{model_name}")
        
        schema_info = []
        for prop in model_details['schema']:
            schema_info.append(f"{prop['name']}: {prop['type']}{prop['required']}")
            if prop['description']:
                schema_info.append(f"  Description: {prop['description']}")
            if prop['example']:
                example = prop['example']
                if isinstance(example, (dict, list)):
                    example = json.dumps(example, indent=2)
                schema_info.append(f"  Example: {example}")
            schema_info.append("")
        
        schema_data.append([
            model_name,
            model_details['description'],
            "\n".join(schema_info).strip()
        ])
    
    schema_headers = [
        "Model Name", "Description", "Schema Details"
    ]
    create_excel_sheet(wb.create_sheet("Schema Details"), "Schema Details", schema_headers, schema_data)
    
    # Add metadata sheet
    meta = wb.create_sheet("Documentation Info")
    meta_data = [
        ["API Documentation Generated On", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["API Version", swagger_data.get('info', {}).get('version', '')],
        ["API Title", swagger_data.get('info', {}).get('title', '')],
        ["API Description", swagger_data.get('info', {}).get('description', '')],
        ["Base URL", swagger_data.get('host', '') + swagger_data.get('basePath', '')],
        ["Generated By", "Swagger to Excel Converter"]
    ]
    
    for row_num, row_data in enumerate(meta_data, 1):
        for col_num, cell_value in enumerate(row_data, 1):
            cell = meta.cell(row=row_num, column=col_num, value=cell_value)
            if row_num == 1:
                cell.font = Font(bold=True)
    
    # Save the workbook
    print(f"Saving to {output_file}...")
    wb.save(output_file)
    print(f"Successfully saved comprehensive API documentation to {output_file}")

if __name__ == "__main__":
    # Configuration
    SWAGGER_URL = "https://api.jobtrees.com/v2/api-docs"
    OUTPUT_FILE = "jobtrees_api_documentation_complete.xlsx"
    
    # Fetch and process the API documentation
    swagger_data = fetch_swagger_docs(SWAGGER_URL)
    if swagger_data:
        parse_swagger_to_excel(swagger_data, OUTPUT_FILE)